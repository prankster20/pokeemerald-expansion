#!/usr/bin/env python3
"""Generate a Second Nature early-game balance workbook.

Edit only the USER CONFIGURATION section for normal use, then run:

    python generate_balance_workbook.py

The script consumes the same generated ``dex.json`` used by romhack-docs.  Run
the docs data generator whenever the decomp changes, then rerun this script.

This is intentionally a transparent balance-analysis engine rather than a
replacement for the ROM's battle code. Unsupported or contextual mechanics are
listed in the workbook instead of being silently guessed.
"""

from __future__ import annotations

import json
import math
import os
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - user-facing dependency check
    raise SystemExit(
        "This tool needs openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc


# =============================================================================
# USER CONFIGURATION
# =============================================================================

# Leave as None to auto-detect either a Windows Downloads/Decomps layout or a
# WSL ~/decomps layout. You may also provide an absolute path.
DEX_JSON: str | None = None

OUTPUT_FILE = "second_nature_roxanne_balance.xlsx"

# Google Sheets export. The spreadsheet URL stays the same on every run; the
# generated tabs are deleted and rebuilt so stale rows/formatting cannot linger.
GOOGLE_SHEETS_ENABLED = True
GOOGLE_SPREADSHEET_ID = "1hdZdmDTvVAbuoer8FHvYYZAVdjp9NGGppLklhBIi62w"
# Normal reruns preserve hand-edited balancing databases. Set this to True
# only when you intentionally want to replace those tabs from dex.json again.
RESET_GOOGLE_EDITOR_TABS = False
GOOGLE_EDITOR_TABS = {
    "Pokemon Editor", "Move Editor", "Nature Editor", "Type Chart", "Boss Editor",
    "Live Calculator", "Threshold Summary", "Evolution Editor", "Learnset Editor",
}
GOOGLE_SERVICE_ACCOUNT_FILE = os.environ.get(
    "SECOND_NATURE_GOOGLE_CREDENTIALS",
    "google-sheets-credentials.json",
)

SCENARIO_NAME = "Pre-Roxanne: encounters through Route 116"
PLAYER_LEVEL = 15
PLAYER_IVS = {"hp": 31, "attack": 31, "defense": 31,
              "spAttack": 31, "spDefense": 31, "speed": 31}
PLAYER_EVS = {"hp": 0, "attack": 0, "defense": 0,
              "spAttack": 0, "spDefense": 0, "speed": 0}

# Nature values:
#   "Neutral"     -> ignore Nature effects for clean species comparisons
#   any real name -> apply that Nature where supported
# PLAYER_NATURE_MODE may be "fixed" or "all". "all" writes one result per
# Pokémon/Nature combination and can produce a very large workbook.
PLAYER_NATURE_MODE = "fixed"
PLAYER_NATURE = "Neutral"

# Ability modes:
#   "first" -> use the first legal ability slot (recommended baseline)
#   "best"  -> try every ability and retain the strongest offensive result
PLAYER_ABILITY_MODE = "first"
PLAYER_ABILITY_OVERRIDES: dict[str, str] = {}

# Legal move policy. Level-up moves are always included. Relearner access adds
# level-1 moves naturally because they are in the level-up list.
INCLUDE_EGG_MOVES = False
INCLUDE_ALL_TEACHABLE_MOVES = False
ACCESSIBLE_TEACHABLE_MOVES: list[str] = []  # e.g. ["MOVE_ROCK_SMASH"]
EXTRA_SPECIES_MOVES: dict[str, list[str]] = {}

# Moves whose listed power is misleading without an unmet prerequisite. Add
# or remove IDs here as your early-game item/move availability changes.
EXCLUDED_PREREQUISITE_MOVES = {
    "MOVE_BELCH",  # fails unless the user has consumed a Berry this battle
}

# Live Sheets resolves evolution eligibility from Evolution Editor instead of
# permanently adding evolved forms to the generated encounter pool.
INCLUDE_LEVEL_EVOLUTIONS = False

# Encounter maps included in this first analysis. Surf is excluded initially
# because it is not available before Roxanne.
ENCOUNTER_MAPS = [
    "MAP_ROUTE101",
    "MAP_ROUTE103",
    "MAP_ROUTE102",
    "MAP_ROUTE104",
    "MAP_PETALBURG_WOODS",
    "MAP_ROUTE116",
]
ALLOWED_ENCOUNTER_METHODS = {"Land", "Sand"}
EXTRA_PLAYER_SPECIES = [
    "SPECIES_TREECKO",
    "SPECIES_TORCHIC",
    "SPECIES_MUDKIP",
]

# The Downloads snapshot still contains vanilla Roxanne, so the starter preset
# below intentionally models the fossil concept. Empty move lists mean "use all
# legal level-up moves at this level". Replace them with the exact trainer moves
# once finalized. Likewise, set ability/item/Nature/EVs to match trainers.party.
BOSS_NAME = "Roxanne (Fossil draft)"
BOSS_TEAM = [
    {
        "species": "SPECIES_TIRTOUGA",
        "level": 12,
        "nature": "Neutral",
        "ability": None,
        "item": None,
        "moves": [],
        "ivs": {"hp": 31, "attack": 31, "defense": 31,
                "spAttack": 31, "spDefense": 31, "speed": 31},
        "evs": {"hp": 0, "attack": 0, "defense": 0,
                "spAttack": 0, "spDefense": 0, "speed": 0},
    },
    {
        "species": "SPECIES_ANORITH",
        "level": 12,
        "nature": "Neutral",
        "ability": None,
        "item": None,
        "moves": [],
        "ivs": {"hp": 31, "attack": 31, "defense": 31,
                "spAttack": 31, "spDefense": 31, "speed": 31},
        "evs": {"hp": 0, "attack": 0, "defense": 0,
                "spAttack": 0, "spDefense": 0, "speed": 0},
    },
    {
        "species": "SPECIES_LILEEP",
        "level": 15,
        "nature": "Neutral",
        "ability": None,
        "item": None,
        "moves": [],
        "ivs": {"hp": 31, "attack": 31, "defense": 31,
                "spAttack": 31, "spDefense": 31, "speed": 31},
        "evs": {"hp": 0, "attack": 0, "defense": 0,
                "spAttack": 0, "spDefense": 0, "speed": 0},
    },
]
BOSS_TARGET_INDEX = 2  # Lileep
BOSS_ATTACKER_INDEX = 2

# Field assumptions. Supported weather: None, Sun, Rain, Sand, Snow.
# Supported terrain: None, Electric, Grassy, Psychic, Misty.
WEATHER = None
TERRAIN = None
PLAYER_STATUS = None  # None, Burn, Poison, Paralysis
BOSS_STATUS = None
PLAYER_AT_FULL_HP = True
BOSS_AT_FULL_HP = True
PLAYER_FIRST_TURN = True

# Ranking by raw average damage favors inaccurate moves. Expected damage folds
# accuracy into the comparison and is the safer learnset-balancing default.
BEST_MOVE_METRIC = "expected"  # "expected", "average", or "minimum"


# =============================================================================
# ENGINE DATA
# =============================================================================

STAT_KEYS = ("hp", "attack", "defense", "spAttack", "spDefense", "speed")
DISPLAY_STATS = {
    "hp": "HP", "attack": "Atk", "defense": "Def",
    "spAttack": "SpA", "spDefense": "SpD", "speed": "Spe",
}

TYPE_CHART: dict[str, dict[str, float]] = {
    "Normal": {"Rock": .5, "Ghost": 0, "Steel": .5},
    "Fire": {"Fire": .5, "Water": .5, "Grass": 2, "Ice": 2,
             "Bug": 2, "Rock": .5, "Dragon": .5, "Steel": 2},
    "Water": {"Fire": 2, "Water": .5, "Grass": .5, "Ground": 2,
              "Rock": 2, "Dragon": .5},
    "Electric": {"Water": 2, "Electric": .5, "Grass": .5, "Ground": 0,
                 "Flying": 2, "Dragon": .5},
    "Grass": {"Fire": .5, "Water": 2, "Grass": .5, "Poison": .5,
              "Ground": 2, "Flying": .5, "Bug": .5, "Rock": 2,
              "Dragon": .5, "Steel": .5},
    "Ice": {"Fire": .5, "Water": .5, "Grass": 2, "Ice": .5,
            "Ground": 2, "Flying": 2, "Dragon": 2, "Steel": .5},
    "Fighting": {"Normal": 2, "Ice": 2, "Poison": .5, "Flying": .5,
                 "Psychic": .5, "Bug": .5, "Rock": 2, "Ghost": 0,
                 "Dark": 2, "Steel": 2, "Fairy": .5},
    "Poison": {"Grass": 2, "Poison": .5, "Ground": .5, "Rock": .5,
               "Ghost": .5, "Steel": 0, "Fairy": 2},
    "Ground": {"Fire": 2, "Electric": 2, "Grass": .5, "Poison": 2,
               "Flying": 0, "Bug": .5, "Rock": 2, "Steel": 2},
    "Flying": {"Electric": .5, "Grass": 2, "Fighting": 2, "Bug": 2,
               "Rock": .5, "Steel": .5},
    "Psychic": {"Fighting": 2, "Poison": 2, "Psychic": .5,
                "Dark": 0, "Steel": .5},
    "Bug": {"Fire": .5, "Grass": 2, "Fighting": .5, "Poison": .5,
            "Flying": .5, "Psychic": 2, "Ghost": .5, "Dark": 2,
            "Steel": .5, "Fairy": .5},
    "Rock": {"Fire": 2, "Ice": 2, "Fighting": .5, "Ground": .5,
             "Flying": 2, "Bug": 2, "Steel": .5},
    "Ghost": {"Normal": 0, "Psychic": 2, "Ghost": 2, "Dark": .5},
    "Dragon": {"Dragon": 2, "Steel": .5, "Fairy": 0},
    "Dark": {"Fighting": .5, "Psychic": 2, "Ghost": 2,
             "Dark": .5, "Fairy": .5},
    "Steel": {"Fire": .5, "Water": .5, "Electric": .5, "Ice": 2,
              "Rock": 2, "Steel": .5, "Fairy": 2},
    "Fairy": {"Fire": .5, "Fighting": 2, "Poison": .5,
              "Dragon": 2, "Dark": 2, "Steel": .5},
}

OLD_PHYSICAL_TYPES = {
    "Normal", "Fighting", "Flying", "Poison", "Ground", "Rock",
    "Bug", "Ghost", "Steel",
}

# These move-family lists support common ability modifiers. Add custom moves as
# needed; omitted tags are disclosed on the Coverage sheet.
PUNCH_MOVES = {
    "MOVE_BULLET_PUNCH", "MOVE_COMET_PUNCH", "MOVE_DIZZY_PUNCH",
    "MOVE_DRAIN_PUNCH", "MOVE_DYNAMIC_PUNCH", "MOVE_FIRE_PUNCH",
    "MOVE_FOCUS_PUNCH", "MOVE_HAMMER_ARM", "MOVE_ICE_HAMMER",
    "MOVE_ICE_PUNCH", "MOVE_MACH_PUNCH", "MOVE_MEGA_PUNCH",
    "MOVE_METEOR_MASH", "MOVE_PLASMA_FISTS", "MOVE_POWER_UP_PUNCH",
    "MOVE_SHADOW_PUNCH", "MOVE_SKY_UPPERCUT", "MOVE_SURGING_STRIKES",
    "MOVE_THUNDER_PUNCH", "MOVE_WICKED_BLOW",
}
BITING_MOVES = {
    "MOVE_BITE", "MOVE_CRUNCH", "MOVE_FIRE_FANG", "MOVE_FISHIOUS_REND",
    "MOVE_HYPER_FANG", "MOVE_ICE_FANG", "MOVE_JAW_LOCK", "MOVE_POISON_FANG",
    "MOVE_PSYCHIC_FANGS", "MOVE_THUNDER_FANG",
}
PULSE_MOVES = {
    "MOVE_AURA_SPHERE", "MOVE_DARK_PULSE", "MOVE_DRAGON_PULSE",
    "MOVE_HEAL_PULSE", "MOVE_ORIGIN_PULSE", "MOVE_TERRAIN_PULSE",
    "MOVE_WATER_PULSE",
}
SLICING_MOVES = {
    "MOVE_AERIAL_ACE", "MOVE_AIR_CUTTER", "MOVE_AIR_SLASH", "MOVE_CEASELESS_EDGE",
    "MOVE_CROSS_POISON", "MOVE_CUT", "MOVE_FURY_CUTTER", "MOVE_LEAF_BLADE",
    "MOVE_NIGHT_SLASH", "MOVE_PSYCHO_CUT", "MOVE_RAZOR_LEAF", "MOVE_RAZOR_SHELL",
    "MOVE_SACRED_SWORD", "MOVE_SLASH", "MOVE_SOLAR_BLADE", "MOVE_X_SCISSOR",
}
SOUND_MOVES = {
    "MOVE_BOOMBURST", "MOVE_BUG_BUZZ", "MOVE_CHATTER", "MOVE_ECHOED_VOICE",
    "MOVE_HYPER_VOICE", "MOVE_OVERDRIVE", "MOVE_ROUND", "MOVE_SNARL",
    "MOVE_SNORE", "MOVE_SPARKLING_ARIA", "MOVE_UPROAR",
}
RECOIL_MOVES = {
    "MOVE_BRAVE_BIRD", "MOVE_DOUBLE_EDGE", "MOVE_FLARE_BLITZ", "MOVE_HEAD_CHARGE",
    "MOVE_HEAD_SMASH", "MOVE_LIGHT_OF_RUIN", "MOVE_SUBMISSION", "MOVE_TAKE_DOWN",
    "MOVE_VOLT_TACKLE", "MOVE_WAVE_CRASH", "MOVE_WILD_CHARGE", "MOVE_WOOD_HAMMER",
}
CONTACT_EXCEPTIONS = {
    # A deliberately small explicit list. The decomp move flags should be added
    # to dex.json later for perfect Tough Claws / Fluffy support.
}


@dataclass(frozen=True)
class Combatant:
    species_id: str
    name: str
    level: int
    types: tuple[str, ...]
    base_stats: dict[str, int]
    stats: dict[str, int]
    nature: str
    ability_id: str | None
    ability_name: str
    item: str | None
    status: str | None
    full_hp: bool


@dataclass(frozen=True)
class DamageResult:
    move_id: str
    move_name: str
    category: str
    move_type: str
    power: int
    accuracy: int
    effectiveness: float
    stab: float
    rolls: tuple[int, ...]
    expected: float
    notes: tuple[str, ...]

    @property
    def minimum(self) -> int:
        return min(self.rolls) if self.rolls else 0

    @property
    def maximum(self) -> int:
        return max(self.rolls) if self.rolls else 0

    @property
    def average(self) -> float:
        return sum(self.rolls) / len(self.rolls) if self.rolls else 0


def find_dex_json() -> Path:
    candidates: list[Path] = []
    if DEX_JSON:
        candidates.append(Path(DEX_JSON).expanduser())
    env_path = os.environ.get("SECOND_NATURE_DEX_JSON")
    if env_path:
        candidates.append(Path(env_path).expanduser())
    candidates.extend([
        Path.home() / "decomps/romhack-docs/public/data/dex.json",
        Path.home() / "Downloads/Decomps/romhack-docs/public/data/dex.json",
        Path("C:/Users/prana/Downloads/Decomps/romhack-docs/public/data/dex.json"),
        Path(__file__).resolve().parents[1] / "romhack-docs/public/data/dex.json",
        Path(__file__).resolve().parents[2] / "romhack-docs/public/data/dex.json",
    ])
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()
    searched = "\n".join(f"  - {path}" for path in candidates)
    raise FileNotFoundError(
        "Could not locate romhack-docs/public/data/dex.json. Set DEX_JSON at the "
        f"top of this script. Searched:\n{searched}"
    )


def norm_id(value: str | None, prefix: str) -> str | None:
    if not value:
        return None
    value = value.strip()
    if value.upper().startswith(prefix):
        return value.upper()
    return prefix + value.upper().replace("-", "_").replace(" ", "_")


def zero_spread() -> dict[str, int]:
    return {key: 0 for key in STAT_KEYS}


def complete_spread(value: dict[str, int] | None, default: int = 0) -> dict[str, int]:
    value = value or {}
    return {key: int(value.get(key, default)) for key in STAT_KEYS}


def type_effectiveness(move_type: str, target_types: Iterable[str]) -> float:
    result = 1.0
    chart = TYPE_CHART.get(move_type, {})
    for target_type in target_types:
        result *= chart.get(target_type, 1.0)
    return result


def nature_stat_multipliers(
    nature_name: str,
    base_stats: dict[str, int],
    nature_by_name: dict[str, dict[str, Any]],
    context: dict[str, Any],
) -> tuple[dict[str, float], list[str]]:
    multipliers = {key: 1.0 for key in STAT_KEYS}
    notes: list[str] = []
    if not nature_name or nature_name.lower() == "neutral":
        return multipliers, notes

    nature = nature_by_name.get(nature_name.lower())
    if not nature:
        notes.append(f"Unknown Nature '{nature_name}'; treated as neutral")
        return multipliers, notes

    stat_map = {
        "Attack": "attack", "Defense": "defense", "Sp Attack": "spAttack",
        "SpAttack": "spAttack", "Sp Defense": "spDefense",
        "SpDefense": "spDefense", "Speed": "speed", "None": None,
    }
    up = stat_map.get(str(nature.get("statUp", "")))
    down = stat_map.get(str(nature.get("statDown", "")))
    name = nature["name"]
    # These Natures override the legacy statUp/statDown metadata in the ROM's
    # GetNatureStatModifierPercent.  The metadata remains useful for summary
    # coloring, but must never be stacked with the custom modifier here.
    custom_stat_natures = {
        "Humble", "Finicky", "Vain", "Noble", "Quirky", "Stoic",
        "Dreamy", "Lazy", "Anxious", "Soft-Hearted", "Proud", "Loyal",
        "Youthful",
    }
    if name not in custom_stat_natures and up != down:
        if up:
            multipliers[up] = 1.15
        if down:
            multipliers[down] = .85

    if name == "Quirky":
        for key in STAT_KEYS[1:]:
            multipliers[key] *= context.get("quirky_multiplier", 1.03)
        notes.append("Quirky modeled at configurable midpoint 3%")
    elif name == "Finicky":
        multipliers["speed"] *= 1.20
    elif name == "Noble":
        multipliers["defense"] *= 1.05
    elif name == "Dreamy":
        multipliers["speed"] *= .90
    elif name == "Lazy":
        multipliers["speed"] *= .90
    elif name == "Stoic":
        multipliers["speed"] *= .95
    elif name == "Humble":
        for key in STAT_KEYS[1:]:
            multipliers[key] *= 1.05
    elif name == "Proud":
        ordered = sorted(STAT_KEYS[1:], key=lambda key: (base_stats[key], -STAT_KEYS.index(key)))
        multipliers[ordered[-1]] *= 1.10
        multipliers[ordered[0]] *= .90
        notes.append(f"Proud boosts {DISPLAY_STATS[ordered[-1]]}, lowers {DISPLAY_STATS[ordered[0]]}")
    elif name == "Soft-Hearted":
        multipliers["spDefense"] *= 1.10
    elif name == "Vain":
        if not context.get("vain_broken", False):
            for key in STAT_KEYS[1:]:
                multipliers[key] *= 1.05
        else:
            notes.append("Vain modeled after its boost has broken")
    elif name == "Anxious":
        multipliers["speed"] *= .80 if context.get("full_hp", True) else 1.20
    elif name == "Loyal":
        gained = max(0, context.get("level", 1) - context.get("met_level", context.get("level", 1)))
        bonus = min(15, gained // 3) / 100
        multipliers["attack"] *= 1 + bonus
        multipliers["spAttack"] *= 1 + bonus
        notes.append(f"Loyal modeled with {bonus:.0%} bonus from met level")
    elif name == "Youthful" and context.get("past_evolution_level", False):
        multipliers["attack"] *= 1.20
        multipliers["spAttack"] *= 1.20
        multipliers["speed"] *= 1.20
    return multipliers, notes


def calculate_stats(
    pokemon: dict[str, Any],
    level: int,
    ivs: dict[str, int],
    evs: dict[str, int],
    nature_name: str,
    nature_by_name: dict[str, dict[str, Any]],
    context: dict[str, Any],
) -> tuple[dict[str, int], list[str]]:
    base = pokemon["stats"]
    nature_multipliers, notes = nature_stat_multipliers(
        nature_name, base, nature_by_name, {**context, "level": level}
    )
    stats: dict[str, int] = {}
    stats["hp"] = math.floor(((2 * base["hp"] + ivs["hp"] + evs["hp"] // 4) * level) / 100) + level + 10
    for key in STAT_KEYS[1:]:
        raw = math.floor(((2 * base[key] + ivs[key] + evs[key] // 4) * level) / 100) + 5
        stats[key] = math.floor(raw * nature_multipliers[key])
    return stats, notes


def ability_name(ability_id: str | None, ability_by_id: dict[str, dict[str, Any]]) -> str:
    if not ability_id:
        return "None"
    return ability_by_id.get(ability_id, {}).get("name", ability_id.removeprefix("ABILITY_").title())


def make_combatant(
    pokemon: dict[str, Any],
    level: int,
    ivs: dict[str, int],
    evs: dict[str, int],
    nature: str,
    ability_id: str | None,
    item: str | None,
    status: str | None,
    full_hp: bool,
    nature_by_name: dict[str, dict[str, Any]],
    ability_by_id: dict[str, dict[str, Any]],
    met_level: int | None = None,
) -> tuple[Combatant, list[str]]:
    stats, notes = calculate_stats(
        pokemon, level, complete_spread(ivs, 31), complete_spread(evs), nature,
        nature_by_name,
        {"full_hp": full_hp, "met_level": met_level if met_level is not None else level,
         "quirky_multiplier": 1.03},
    )
    return Combatant(
        species_id=pokemon["id"], name=pokemon["name"], level=level,
        types=tuple(pokemon["types"]), base_stats=pokemon["stats"], stats=stats,
        nature=nature, ability_id=ability_id, ability_name=ability_name(ability_id, ability_by_id),
        item=item, status=status, full_hp=full_hp,
    ), notes


def legal_moves(
    pokemon: dict[str, Any], level: int, move_by_id: dict[str, dict[str, Any]],
    nature: str = "Neutral", explicit_moves: Iterable[str] | None = None,
) -> list[dict[str, Any]]:
    if explicit_moves:
        ids = [norm_id(move, "MOVE_") for move in explicit_moves]
    else:
        effective_level = level + (1 if nature == "Prodigious" else 0)
        ids = [entry["move"] for entry in pokemon["learnsets"]["levelUp"]
               if int(entry["level"]) <= effective_level]
        if INCLUDE_EGG_MOVES or nature == "Scholarly":
            ids.extend(pokemon["learnsets"]["egg"])
        if INCLUDE_ALL_TEACHABLE_MOVES:
            ids.extend(pokemon["learnsets"]["teachable"])
        accessible = {norm_id(move, "MOVE_") for move in ACCESSIBLE_TEACHABLE_MOVES}
        ids.extend(move for move in pokemon["learnsets"]["teachable"] if move in accessible)
        ids.extend(norm_id(move, "MOVE_") for move in EXTRA_SPECIES_MOVES.get(pokemon["id"], []))

    seen: set[str] = set()
    result = []
    for move_id in ids:
        if move_id in EXCLUDED_PREREQUISITE_MOVES:
            continue
        if move_id and move_id not in seen and move_id in move_by_id:
            result.append(move_by_id[move_id])
            seen.add(move_id)
    return result


def move_is_contact(move: dict[str, Any]) -> bool:
    # The docs snapshot currently lacks move flags. This conservative heuristic
    # is only used for Tough Claws / Fluffy and is disclosed in Coverage.
    if move["id"] in CONTACT_EXCEPTIONS:
        return False
    if move["category"] != "Physical":
        return False
    noncontact_words = ("Rock", "Stone", "Earthquake", "Bulldoze", "Bone", "Shard", "Shot")
    return not any(word in move["name"] for word in noncontact_words)


def category_for(move: dict[str, Any], attacker: Combatant) -> tuple[str, list[str]]:
    category = move["category"]
    notes: list[str] = []
    if attacker.nature == "Old-Fashioned" and category in {"Physical", "Special"}:
        category = "Physical" if move["type"] in OLD_PHYSICAL_TYPES else "Special"
        notes.append(f"Old-Fashioned changes category to {category}")
    return category, notes


def apply_ability_immunities(
    effectiveness: float, move: dict[str, Any], defender: Combatant, notes: list[str]
) -> float:
    ability = defender.ability_id
    move_type = move["type"]
    immunities = {
        "ABILITY_LEVITATE": "Ground",
        "ABILITY_FLASH_FIRE": "Fire",
        "ABILITY_WATER_ABSORB": "Water",
        "ABILITY_DRY_SKIN": "Water",
        "ABILITY_STORM_DRAIN": "Water",
        "ABILITY_VOLT_ABSORB": "Electric",
        "ABILITY_LIGHTNING_ROD": "Electric",
        "ABILITY_MOTOR_DRIVE": "Electric",
        "ABILITY_SAP_SIPPER": "Grass",
        "ABILITY_EARTH_EATER": "Ground",
    }
    if ability in immunities and move_type == immunities[ability]:
        notes.append(f"{defender.ability_name} grants immunity")
        return 0
    return effectiveness


def offensive_stat(attacker: Combatant, category: str, move: dict[str, Any], notes: list[str]) -> int:
    key = "attack" if category == "Physical" else "spAttack"
    value = attacker.stats[key]
    ability = attacker.ability_id
    if category == "Physical" and ability in {"ABILITY_HUGE_POWER", "ABILITY_PURE_POWER"}:
        value *= 2
        notes.append(f"{attacker.ability_name} doubles Attack")
    if category == "Physical" and ability == "ABILITY_HUSTLE":
        value = math.floor(value * 1.5)
        notes.append("Hustle Attack boost")
    if category == "Physical" and ability == "ABILITY_GUTS" and attacker.status:
        value = math.floor(value * 1.5)
        notes.append("Guts activated")
    if category == "Physical" and ability == "ABILITY_GORILLA_TACTICS":
        value = math.floor(value * 1.5)
        notes.append("Gorilla Tactics activated")
    if category == "Special" and ability == "ABILITY_SOLAR_POWER" and WEATHER == "Sun":
        value = math.floor(value * 1.5)
        notes.append("Solar Power activated")
    return max(1, value)


def defensive_stat(defender: Combatant, category: str, notes: list[str]) -> int:
    key = "defense" if category == "Physical" else "spDefense"
    value = defender.stats[key]
    if category == "Physical" and defender.ability_id == "ABILITY_FUR_COAT":
        value *= 2
        notes.append("Fur Coat doubles Defense")
    return max(1, value)


def damage_modifiers(
    attacker: Combatant, defender: Combatant, move: dict[str, Any],
    category: str, effectiveness: float, notes: list[str], first_turn: bool,
) -> tuple[float, float, int]:
    modifier = 1.0
    accuracy = int(move.get("accuracy") or 100)
    move_type = move["type"]
    ability = attacker.ability_id
    defending_ability = defender.ability_id

    stab = 1.0
    if move_type in attacker.types:
        stab = 2.0 if ability == "ABILITY_ADAPTABILITY" else 1.5

    if WEATHER == "Sun":
        if move_type == "Fire": modifier *= 1.5
        elif move_type == "Water": modifier *= .5
    elif WEATHER == "Rain":
        if move_type == "Water": modifier *= 1.5
        elif move_type == "Fire": modifier *= .5

    if TERRAIN == "Electric" and move_type == "Electric": modifier *= 1.3
    elif TERRAIN == "Grassy" and move_type == "Grass": modifier *= 1.3
    elif TERRAIN == "Psychic" and move_type == "Psychic": modifier *= 1.3
    elif TERRAIN == "Misty" and move_type == "Dragon": modifier *= .5

    power = int(move.get("power") or 0)
    if ability == "ABILITY_TECHNICIAN" and power <= 60:
        modifier *= 1.5; notes.append("Technician")
    if ability == "ABILITY_IRON_FIST" and move["id"] in PUNCH_MOVES:
        modifier *= 1.2; notes.append("Iron Fist")
    if ability == "ABILITY_STRONG_JAW" and move["id"] in BITING_MOVES:
        modifier *= 1.5; notes.append("Strong Jaw")
    if ability == "ABILITY_MEGA_LAUNCHER" and move["id"] in PULSE_MOVES:
        modifier *= 1.5; notes.append("Mega Launcher")
    if ability == "ABILITY_SHARPNESS" and move["id"] in SLICING_MOVES:
        modifier *= 1.5; notes.append("Sharpness")
    if ability == "ABILITY_PUNK_ROCK" and move["id"] in SOUND_MOVES:
        modifier *= 1.3; notes.append("Punk Rock")
    if ability == "ABILITY_RECKLESS" and move["id"] in RECOIL_MOVES:
        modifier *= 1.2; notes.append("Reckless")
    if ability == "ABILITY_TOUGH_CLAWS" and move_is_contact(move):
        modifier *= 1.3; notes.append("Tough Claws (contact inferred)")
    if ability == "ABILITY_STEELWORKER" and move_type == "Steel": modifier *= 1.5
    if ability == "ABILITY_TRANSISTOR" and move_type == "Electric": modifier *= 1.3
    if ability == "ABILITY_DRAGONS_MAW" and move_type == "Dragon": modifier *= 1.5
    if ability == "ABILITY_WATER_BUBBLE" and move_type == "Water": modifier *= 2
    if ability == "ABILITY_HUSTLE" and category == "Physical": accuracy = math.floor(accuracy * .8)

    if attacker.nature == "Arrogant" and defender.level > attacker.level:
        modifier *= 1.10; notes.append("Arrogant")
    if attacker.nature == "Relentless":
        modifier *= 1.10; notes.append("Relentless")
    if attacker.nature == "Shortsighted" and first_turn:
        modifier *= 1.20; notes.append("Shortsighted first turn")
    if attacker.nature == "Scholarly":
        # Exact acquisition source is tracked by the caller in a future version.
        notes.append("Scholarly 10% move-source boost not inferred")
    if attacker.nature == "Perfectionist":
        accuracy = 100; notes.append("Perfectionist guarantees accuracy")

    if effectiveness > 1 and defending_ability in {
        "ABILITY_FILTER", "ABILITY_SOLID_ROCK", "ABILITY_PRISM_ARMOR"
    }:
        modifier *= .75; notes.append(defender.ability_name)
    if effectiveness < 1 and effectiveness > 0 and ability == "ABILITY_TINTED_LENS":
        modifier *= 2; notes.append("Tinted Lens")
    if defending_ability == "ABILITY_THICK_FAT" and move_type in {"Fire", "Ice"}:
        modifier *= .5; notes.append("Thick Fat")
    if defending_ability == "ABILITY_HEATPROOF" and move_type == "Fire":
        modifier *= .5; notes.append("Heatproof")
    if defending_ability == "ABILITY_WATER_BUBBLE" and move_type == "Fire":
        modifier *= .5; notes.append("Water Bubble")
    if defending_ability == "ABILITY_DRY_SKIN" and move_type == "Fire":
        modifier *= 1.25; notes.append("Dry Skin")
    if defending_ability == "ABILITY_ICE_SCALES" and category == "Special":
        modifier *= .5; notes.append("Ice Scales")
    if defending_ability == "ABILITY_MULTISCALE" and defender.full_hp:
        modifier *= .5; notes.append("Multiscale at full HP")
    if defending_ability == "ABILITY_FLUFFY":
        if move_is_contact(move): modifier *= .5; notes.append("Fluffy contact reduction (inferred)")
        if move_type == "Fire": modifier *= 2; notes.append("Fluffy Fire weakness")
    if defending_ability == "ABILITY_PUNK_ROCK" and move["id"] in SOUND_MOVES:
        modifier *= .5; notes.append("Punk Rock sound resistance")

    if category == "Physical" and attacker.status == "Burn" and ability != "ABILITY_GUTS":
        modifier *= .5; notes.append("Burn")
    return modifier, stab, max(0, min(100, accuracy))


def calculate_damage(
    attacker: Combatant, defender: Combatant, move: dict[str, Any], first_turn: bool,
) -> DamageResult | None:
    if move.get("category") == "Status" or int(move.get("power") or 0) <= 0:
        return None
    notes: list[str] = []
    category, category_notes = category_for(move, attacker)
    notes.extend(category_notes)
    effectiveness = type_effectiveness(move["type"], defender.types)
    effectiveness = apply_ability_immunities(effectiveness, move, defender, notes)
    if effectiveness == 0:
        return DamageResult(move["id"], move["name"], category, move["type"],
                            int(move["power"]), int(move.get("accuracy") or 100),
                            0, 1, tuple([0] * 16), 0, tuple(notes))

    attack = offensive_stat(attacker, category, move, notes)
    defense = defensive_stat(defender, category, notes)
    modifier, stab, accuracy = damage_modifiers(
        attacker, defender, move, category, effectiveness, notes, first_turn
    )
    base = math.floor(math.floor(math.floor((2 * attacker.level) / 5 + 2)
                                 * int(move["power"]) * attack / defense) / 50) + 2
    rolls = tuple(max(1, math.floor(base * modifier * stab * effectiveness * roll / 100))
                  for roll in range(85, 101))
    expected = (sum(rolls) / len(rolls)) * accuracy / 100
    return DamageResult(move["id"], move["name"], category, move["type"],
                        int(move["power"]), accuracy, effectiveness, stab,
                        rolls, expected, tuple(dict.fromkeys(notes)))


def metric(result: DamageResult) -> float:
    if BEST_MOVE_METRIC == "minimum":
        return result.minimum
    if BEST_MOVE_METRIC == "average":
        return result.average
    return result.expected


def best_result(
    attacker: Combatant, defender: Combatant, moves: Iterable[dict[str, Any]],
    first_turn: bool,
) -> DamageResult | None:
    results = [result for move in moves if (result := calculate_damage(attacker, defender, move, first_turn))]
    return max(results, key=lambda result: (metric(result), result.average, result.maximum), default=None)


def ko_chance(rolls: tuple[int, ...], hp: int, hits: int) -> float:
    if not rolls or hp <= 0:
        return 0
    totals = {0: 1}
    for _ in range(hits):
        next_totals: dict[int, int] = defaultdict(int)
        for total, count in totals.items():
            for roll in rolls:
                next_totals[min(hp, total + roll)] += count
        totals = next_totals
    favorable = sum(count for total, count in totals.items() if total >= hp)
    return favorable / (len(rolls) ** hits)


def ko_label(result: DamageResult | None, hp: int) -> str:
    if not result or result.maximum <= 0:
        return "No damage"
    for hits in range(1, 5):
        chance = ko_chance(result.rolls, hp, hits)
        if chance >= 1:
            return f"Guaranteed {hits}HKO"
        if chance > 0:
            return f"{chance:.0%} {hits}HKO"
    hits = math.ceil(hp / max(.001, result.average))
    return f"~{hits}HKO"


def evolution_level(evolution: dict[str, Any]) -> int | None:
    method = str(evolution.get("method", ""))
    if "LEVEL" not in method or "FRIENDSHIP" in method or "GENDER" in method:
        return None
    raw = str(evolution.get("parameter", ""))
    digits = "".join(character for character in raw if character.isdigit())
    level = int(digits) if digits else 0
    # Parameter 0 is used by condition-driven evolutions (friendship/time,
    # etc.). The compact docs snapshot does not retain those conditions, so
    # never mislabel them as a free Lv. 0 evolution.
    return level if level > 0 else None


def collect_player_species(
    data: dict[str, Any], pokemon_by_id: dict[str, dict[str, Any]]
) -> tuple[list[str], dict[str, set[str]]]:
    locations: dict[str, set[str]] = defaultdict(set)
    allowed_maps = set(ENCOUNTER_MAPS)
    for encounter in data.get("encounters", []):
        if encounter["id"] not in allowed_maps:
            continue
        for method in encounter.get("methods", []):
            if method.get("name") not in ALLOWED_ENCOUNTER_METHODS:
                continue
            for period, slots in method.get("periods", {}).items():
                for slot in slots:
                    species = slot["species"]
                    if species in pokemon_by_id:
                        locations[species].add(f"{encounter['name']} / {method['name']} / {period}")
    for species in EXTRA_PLAYER_SPECIES:
        species_id = norm_id(species, "SPECIES_")
        if species_id in pokemon_by_id:
            locations[species_id].add("Starter")

    if INCLUDE_LEVEL_EVOLUTIONS:
        changed = True
        while changed:
            changed = False
            for species_id in list(locations):
                pokemon = pokemon_by_id[species_id]
                for evolution in pokemon.get("evolutions", []):
                    level = evolution_level(evolution)
                    target = evolution.get("target")
                    if level is not None and level <= PLAYER_LEVEL and target in pokemon_by_id and target not in locations:
                        locations[target] = {f"Evolves from {pokemon['name']} at Lv. {level}"}
                        changed = True
    return sorted(locations, key=lambda species_id: pokemon_by_id[species_id]["number"]), locations


def select_abilities(pokemon: dict[str, Any]) -> list[str | None]:
    override = PLAYER_ABILITY_OVERRIDES.get(pokemon["id"])
    if override:
        return [norm_id(override, "ABILITY_")]
    abilities = pokemon.get("abilities") or [None]
    return abilities if PLAYER_ABILITY_MODE == "best" else [abilities[0]]


def nature_choices(natures: list[dict[str, Any]]) -> list[str]:
    if PLAYER_NATURE_MODE == "all":
        return [nature["name"] for nature in natures]
    return [PLAYER_NATURE]


def style_workbook(workbook: Workbook) -> None:
    navy = "172033"
    blue = "2878B5"
    pale = "EAF2F8"
    white = "FFFFFF"
    gray = "5F6B7A"
    thin = Side(style="thin", color="CBD5E1")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions if sheet.max_row > 1 else None
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color=blue))
        sheet.row_dimensions[1].height = 32
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = Border(bottom=thin)
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[: min(len(column_cells), 250)]]
            width = min(42, max(9, max((len(value) for value in values), default=8) + 2))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        if sheet.title in {"Scenario", "Coverage", "Nature Reference"}:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, 1).font = Font(bold=True, color=gray)
        sheet.auto_filter.ref = sheet.dimensions


def google_cell_value(value: Any) -> Any:
    """Convert openpyxl values into JSON-safe Sheets API values."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    return value


GS_BASE_STAT_COLUMNS = {
    "hp": "E", "attack": "F", "defense": "G",
    "spAttack": "H", "spDefense": "I", "speed": "J",
}
GS_NATURE_COLUMNS = {
    "attack": "B", "defense": "C", "spAttack": "D",
    "spDefense": "E", "speed": "F",
}


def gs_stat_expr(species: str, level: str, nature: str, ev: str, stat: str) -> str:
    base_col = GS_BASE_STAT_COLUMNS[stat]
    base = f"XLOOKUP({species},'Pokemon Editor'!$A:$A,'Pokemon Editor'!${base_col}:${base_col},0)"
    raw = f"FLOOR(((2*{base}+31+FLOOR({ev}/4))*{level})/100)"
    if stat == "hp":
        return f"({raw}+{level}+10)"
    nature_col = GS_NATURE_COLUMNS[stat]
    multiplier = f"XLOOKUP({nature},'Nature Editor'!$A:$A,'Nature Editor'!${nature_col}:${nature_col},1)"
    return f"FLOOR(({raw}+5)*{multiplier})"


def gs_type_effectiveness_expr(move_type: str, defender_species: str) -> str:
    type1 = f"XLOOKUP({defender_species},'Pokemon Editor'!$A:$A,'Pokemon Editor'!$C:$C,\"\")"
    type2 = f"XLOOKUP({defender_species},'Pokemon Editor'!$A:$A,'Pokemon Editor'!$D:$D,\"\")"
    first = f"INDEX('Type Chart'!$B:$S,MATCH({move_type},'Type Chart'!$A:$A,0),MATCH({type1},'Type Chart'!$B$1:$S$1,0))"
    second = f"IF({type2}=\"\",1,INDEX('Type Chart'!$B:$S,MATCH({move_type},'Type Chart'!$A:$A,0),MATCH({type2},'Type Chart'!$B$1:$S$1,0)))"
    return f"({first}*{second})"


def gs_damage_formula(
    attacker_species: str, attacker_level: str, attacker_nature: str,
    attacker_atk_ev: str, attacker_spa_ev: str, move_name: str,
    defender_species: str, defender_level: str, defender_nature: str,
    defender_hp_ev: str, defender_def_ev: str, defender_spd_ev: str,
    metric: str,
) -> str:
    move_type = f"XLOOKUP({move_name},'Move Editor'!$B:$B,'Move Editor'!$C:$C,\"\")"
    category = f"XLOOKUP({move_name},'Move Editor'!$B:$B,'Move Editor'!$D:$D,\"\")"
    power = f"XLOOKUP({move_name},'Move Editor'!$B:$B,'Move Editor'!$E:$E,0)"
    accuracy = f"XLOOKUP({move_name},'Move Editor'!$B:$B,'Move Editor'!$F:$F,100)"
    attack = f"IF(cat=\"Physical\",{gs_stat_expr(attacker_species, attacker_level, attacker_nature, attacker_atk_ev, 'attack')},{gs_stat_expr(attacker_species, attacker_level, attacker_nature, attacker_spa_ev, 'spAttack')})"
    defense = f"IF(cat=\"Physical\",{gs_stat_expr(defender_species, defender_level, defender_nature, defender_def_ev, 'defense')},{gs_stat_expr(defender_species, defender_level, defender_nature, defender_spd_ev, 'spDefense')})"
    atk_type1 = f"XLOOKUP({attacker_species},'Pokemon Editor'!$A:$A,'Pokemon Editor'!$C:$C,\"\")"
    atk_type2 = f"XLOOKUP({attacker_species},'Pokemon Editor'!$A:$A,'Pokemon Editor'!$D:$D,\"\")"
    stab = f"IF(OR(mtype={atk_type1},mtype={atk_type2}),1.5,1)"
    effectiveness = gs_type_effectiveness_expr("mtype", defender_species)
    defender_hp = gs_stat_expr(defender_species, defender_level, defender_nature, defender_hp_ev, "hp")
    expression = {
        "min": "FLOOR(base*modifier*0.85)",
        "max": "FLOOR(base*modifier)",
        "average": "AVERAGE(ARRAYFORMULA(FLOOR(base*modifier*SEQUENCE(16,1,85,1)/100)))",
        "expected": "AVERAGE(ARRAYFORMULA(FLOOR(base*modifier*SEQUENCE(16,1,85,1)/100)))*MIN(1,acc/100)",
        "average_pct": f"AVERAGE(ARRAYFORMULA(FLOOR(base*modifier*SEQUENCE(16,1,85,1)/100)))/{defender_hp}",
    }[metric]
    return (
        f'=IFERROR(LET(mtype,{move_type},cat,{category},pow,{power},acc,{accuracy},'
        f'atk,{attack},def,{defense},stab,{stab},eff,{effectiveness},'
        f'base,IF(OR(cat="Status",pow=0),0,FLOOR(FLOOR(FLOOR((2*{attacker_level}/5+2)*pow*atk/def)/50)+2)),'
        f'modifier,stab*eff,{expression}),0)'
    )


def gs_effective_species_formula(origin: str, level: str) -> str:
    # Two passes cover ordinary three-stage families. REDUCE avoids LET names
    # such as ``stage1`` that Google Sheets can ambiguously parse as A1-style
    # cell references.
    return (
        f'=REDUCE({origin},SEQUENCE(2),LAMBDA(current_species,unused_step,'
        f'IFERROR(INDEX(FILTER(\'Evolution Editor\'!$D$2:$D,'
        f'\'Evolution Editor\'!$A$2:$A=current_species,'
        f'\'Evolution Editor\'!$F$2:$F=TRUE,'
        f'\'Evolution Editor\'!$E$2:$E<={level}),1),current_species)))'
    )


def gs_best_legal_move_formula(
    override: str, species: str, level: str, nature: str,
    atk_ev: str, spa_ev: str,
    defender_species: str, defender_level: str, defender_nature: str,
    defender_hp_ev: str, defender_def_ev: str, defender_spd_ev: str,
) -> str:
    damage = gs_damage_formula(
        species, level, nature, atk_ev, spa_ev, "mv",
        defender_species, defender_level, defender_nature,
        defender_hp_ev, defender_def_ev, defender_spd_ev, "average_pct",
    )[1:]
    return (
        f'=IF({override}<>"",{override},IFERROR(LET('
        f'moves,FILTER(\'Learnset Editor\'!$D$2:$D,'
        f'\'Learnset Editor\'!$A$2:$A={species},'
        f'\'Learnset Editor\'!$C$2:$C<={level}),'
        f'scores,MAP(moves,LAMBDA(mv,{damage})),'
        f'INDEX(moves,MATCH(MAX(scores),scores,0))),""))'
    )


def make_analysis_tabs_live(
    threshold: Any, offense: Any, defense: Any, comparison: Any, matchups: Any,
) -> None:
    """Replace snapshot outputs with formulas over the editable database tabs."""
    boss_target_row = BOSS_TARGET_INDEX + 2
    boss_attacker_row = BOSS_ATTACKER_INDEX + 2

    # Threshold Summary is the central editable roster: B/D and L:T are inputs.
    for row in range(2, threshold.max_row + 1):
        threshold.cell(row, 11, gs_effective_species_formula(f"$Y{row}", f"$L{row}"))
        threshold.cell(row, 1, f"=XLOOKUP(K{row},'Pokemon Editor'!$A:$A,'Pokemon Editor'!$B:$B,K{row})")
        sid = f"$K{row}"
        level = f"$L{row}"
        nature = f"$M{row}"
        move = f"$B{row}"
        evs = {"hp": f"$O{row}", "attack": f"$P{row}", "defense": f"$Q{row}",
               "spAttack": f"$R{row}", "spDefense": f"$S{row}", "speed": f"$T{row}"}
        bsid = f"'Boss Editor'!$C${boss_target_row}"
        blvl = f"'Boss Editor'!$D${boss_target_row}"
        bnat = f"'Boss Editor'!$E${boss_target_row}"
        bevs = {"hp": f"'Boss Editor'!$G${boss_target_row}",
                "attack": f"'Boss Editor'!$H${boss_target_row}",
                "defense": f"'Boss Editor'!$I${boss_target_row}",
                "spAttack": f"'Boss Editor'!$J${boss_target_row}",
                "spDefense": f"'Boss Editor'!$K${boss_target_row}",
                "speed": f"'Boss Editor'!$L${boss_target_row}"}
        asid = f"'Boss Editor'!$C${boss_attacker_row}"
        alvl = f"'Boss Editor'!$D${boss_attacker_row}"
        anat = f"'Boss Editor'!$E${boss_attacker_row}"
        aevs = {"hp": f"'Boss Editor'!$G${boss_attacker_row}",
                "attack": f"'Boss Editor'!$H${boss_attacker_row}",
                "defense": f"'Boss Editor'!$I${boss_attacker_row}",
                "spAttack": f"'Boss Editor'!$J${boss_attacker_row}",
                "spDefense": f"'Boss Editor'!$K${boss_attacker_row}",
                "speed": f"'Boss Editor'!$L${boss_attacker_row}"}
        # Move Comparison already evaluates every legal move against the live
        # target. Reuse those results instead of duplicating the full damage
        # engine inside one very large MAP/LAMBDA formula.
        threshold.cell(row, 2, (
            f'=IF($Z{row}<>"",$Z{row},IFERROR(INDEX(FILTER('
            f'\'Move Comparison\'!$D$2:$D,'
            f'\'Move Comparison\'!$Q$2:$Q=$Y{row},'
            f'\'Move Comparison\'!$M$2:$M=MAXIFS('
            f'\'Move Comparison\'!$M$2:$M,'
            f'\'Move Comparison\'!$Q$2:$Q,$Y{row}),'
            f'\'Move Comparison\'!$D$2:$D<>""),1),""))'
        ))
        threshold.cell(row, 3, gs_damage_formula(sid, level, nature, evs["attack"], evs["spAttack"], move,
                                                  bsid, blvl, bnat, bevs["hp"], bevs["defense"], bevs["spDefense"], "average_pct"))
        threshold.cell(row, 5, gs_damage_formula(asid, alvl, anat, aevs["attack"], aevs["spAttack"], f"$D{row}",
                                                  sid, level, nature, evs["hp"], evs["defense"], evs["spDefense"], "average_pct"))
        threshold.cell(row, 6, f"={gs_stat_expr(sid, level, nature, evs['speed'], 'speed')}")
        threshold.cell(row, 7, f"={gs_stat_expr(asid, alvl, anat, aevs['speed'], 'speed')}")
        threshold.cell(row, 8, f'=IF(F{row}>G{row},"Yes","No")')
        threshold.cell(row, 9, f"=C{row}-E{row}")
        threshold.cell(row, 10, f'=IFS(AND(C{row}>=0.5,E{row}<0.5),"Dominant",AND(C{row}>=0.35,E{row}<0.75),"Strong",AND(C{row}>=0.25,E{row}<1),"Viable",AND(C{row}>=0.15,E{row}<1.25),"Risky",TRUE,"Poor")')
        threshold.cell(row, 21, gs_damage_formula(sid, level, nature, evs["attack"], evs["spAttack"], move,
                                                   bsid, blvl, bnat, bevs["hp"], bevs["defense"], bevs["spDefense"], "min"))
        threshold.cell(row, 22, gs_damage_formula(sid, level, nature, evs["attack"], evs["spAttack"], move,
                                                   bsid, blvl, bnat, bevs["hp"], bevs["defense"], bevs["spDefense"], "max"))
        threshold.cell(row, 23, gs_damage_formula(asid, alvl, anat, aevs["attack"], aevs["spAttack"], f"$D{row}",
                                                   sid, level, nature, evs["hp"], evs["defense"], evs["spDefense"], "min"))
        threshold.cell(row, 24, gs_damage_formula(asid, alvl, anat, aevs["attack"], aevs["spAttack"], f"$D{row}",
                                                   sid, level, nature, evs["hp"], evs["defense"], evs["spDefense"], "max"))
        for column in (4, 12, 13, 14, 15, 16, 17, 18, 19, 20, 26):
            threshold.cell(row, column).fill = PatternFill("solid", fgColor="FFF2CC")

    # Player-vs-boss and boss-vs-player are readable projections of Threshold.
    for row in range(2, offense.max_row + 1):
        sid = f"$R{row}"
        lookup = lambda col: f"XLOOKUP({sid},'Threshold Summary'!$Y:$Y,'Threshold Summary'!${col}:${col},\"\")"
        offense.cell(row, 1, f"={lookup('A')}")
        offense.cell(row, 3, f"={lookup('L')}")
        offense.cell(row, 4, f"={lookup('M')}")
        offense.cell(row, 5, f"={lookup('N')}")
        offense.cell(row, 6, f"={lookup('B')}")
        offense.cell(row, 7, f"=XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$C:$C,\"\")")
        offense.cell(row, 8, f"=XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$D:$D,\"\")")
        offense.cell(row, 9, f"=XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$E:$E,0)")
        offense.cell(row, 10, f"=XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$F:$F,0)")
        offense.cell(row, 11, f"={lookup('U')}")
        offense.cell(row, 12, f"={lookup('V')}")
        offense.cell(row, 13, f"=AVERAGE(K{row}:L{row})")
        offense.cell(row, 14, f"=M{row}*J{row}/100")
        offense.cell(row, 15, f"={lookup('C')}")
        offense.cell(row, 16, f'=IF(O{row}>=1,"OHKO",IF(O{row}>=0.5,"2HKO",IF(O{row}>=0.333,"3HKO",IF(O{row}>=0.25,"4HKO","5HKO+"))))')
        offense.cell(row, 17, "Live from Threshold Summary and editor tabs")

    for row in range(2, defense.max_row + 1):
        sid = f"$Q{row}"
        lookup = lambda col: f"XLOOKUP({sid},'Threshold Summary'!$Y:$Y,'Threshold Summary'!${col}:${col},\"\")"
        effective_sid = lookup('K')
        defense.cell(row, 1, f"={lookup('A')}")
        defense.cell(row, 2, f"={lookup('M')}")
        defense.cell(row, 3, f"={lookup('N')}")
        defense.cell(row, 4, f"={gs_stat_expr(effective_sid, lookup('L'), lookup('M'), lookup('O'), 'hp')}")
        defense.cell(row, 6, f"={lookup('D')}")
        defense.cell(row, 7, f"=XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$C:$C,\"\")")
        defense.cell(row, 8, f"=XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$D:$D,\"\")")
        defense.cell(row, 9, f"={lookup('W')}")
        defense.cell(row, 10, f"={lookup('X')}")
        defense.cell(row, 11, f"=AVERAGE(I{row}:J{row})")
        defense.cell(row, 12, f"=K{row}*XLOOKUP(F{row},'Move Editor'!$B:$B,'Move Editor'!$F:$F,100)/100")
        defense.cell(row, 13, f"={lookup('E')}")
        defense.cell(row, 14, f'=IF(M{row}>=1,"OHKO",IF(M{row}>=0.5,"2HKO",IF(M{row}>=0.333,"3HKO",IF(M{row}>=0.25,"4HKO","5HKO+"))))')
        defense.cell(row, 15, f"={lookup('H')}")
        defense.cell(row, 16, "Live from Threshold Summary and editor tabs")

    # Every move-comparison row recalculates from the selected player's inputs.
    for row in range(2, comparison.max_row + 1):
        origin_sid = f"$Q{row}"
        lookup = lambda col: f"XLOOKUP({origin_sid},'Threshold Summary'!$Y:$Y,'Threshold Summary'!${col}:${col},\"\")"
        sid = lookup('K')
        comparison.cell(row, 1, f"={lookup('A')}")
        comparison.cell(row, 4, (
            f'=IFERROR(INDEX(FILTER(\'Learnset Editor\'!$D$2:$D,'
            f'\'Learnset Editor\'!$A$2:$A={sid},'
            f'\'Learnset Editor\'!$C$2:$C<={lookup("L")}),$R{row}),"")'
        ))
        comparison.cell(row, 2, f"={lookup('M')}")
        comparison.cell(row, 3, f"={lookup('N')}")
        comparison.cell(row, 5, f"=XLOOKUP(D{row},'Move Editor'!$B:$B,'Move Editor'!$C:$C,\"\")")
        comparison.cell(row, 6, f"=XLOOKUP(D{row},'Move Editor'!$B:$B,'Move Editor'!$D:$D,\"\")")
        comparison.cell(row, 7, f"=XLOOKUP(D{row},'Move Editor'!$B:$B,'Move Editor'!$E:$E,0)")
        comparison.cell(row, 8, f"=XLOOKUP(D{row},'Move Editor'!$B:$B,'Move Editor'!$F:$F,0)")
        comparison.cell(row, 9, f"={gs_type_effectiveness_expr(f'E{row}', f"'Boss Editor'!$C${boss_target_row}")}")
        args = (sid, lookup('L'), lookup('M'), lookup('P'), lookup('R'), f"$D{row}",
                f"'Boss Editor'!$C${boss_target_row}", f"'Boss Editor'!$D${boss_target_row}",
                f"'Boss Editor'!$E${boss_target_row}", f"'Boss Editor'!$G${boss_target_row}",
                f"'Boss Editor'!$I${boss_target_row}", f"'Boss Editor'!$K${boss_target_row}")
        comparison.cell(row, 10, gs_damage_formula(*args, "min"))
        comparison.cell(row, 11, gs_damage_formula(*args, "max"))
        comparison.cell(row, 12, gs_damage_formula(*args, "average"))
        comparison.cell(row, 13, gs_damage_formula(*args, "expected"))
        comparison.cell(row, 14, gs_damage_formula(*args, "average_pct"))
        comparison.cell(row, 15, f'=IF(N{row}>=1,"OHKO",IF(N{row}>=0.5,"2HKO",IF(N{row}>=0.333,"3HKO",IF(N{row}>=0.25,"4HKO","5HKO+"))))')
        comparison.cell(row, 16, "Live from editable databases")

    for row in range(2, matchups.max_row + 1):
        origin_sid = f"$K{row}"
        bsid = f"$L{row}"
        plook = lambda col: f"XLOOKUP({origin_sid},'Threshold Summary'!$Y:$Y,'Threshold Summary'!${col}:${col},\"\")"
        sid = plook('K')
        matchups.cell(row, 1, f"={plook('A')}")
        boss_lookup = lambda col: f"XLOOKUP({bsid},'Boss Editor'!$C:$C,'Boss Editor'!${col}:${col},\"\")"
        matchups.cell(row, 3, f"={boss_lookup('D')}")
        matchups.cell(row, 4, f"={plook('B')}")
        matchups.cell(row, 5, f"=XLOOKUP(D{row},'Move Editor'!$B:$B,'Move Editor'!$C:$C,\"\")")
        args = (sid, plook('L'), plook('M'), plook('P'), plook('R'), f"$D{row}",
                bsid, boss_lookup('D'), boss_lookup('E'), boss_lookup('G'), boss_lookup('I'), boss_lookup('K'))
        matchups.cell(row, 6, gs_damage_formula(*args, "min"))
        matchups.cell(row, 7, gs_damage_formula(*args, "max"))
        matchups.cell(row, 8, gs_damage_formula(*args, "average"))
        matchups.cell(row, 9, gs_damage_formula(*args, "average_pct"))
        matchups.cell(row, 10, f'=IF(I{row}>=1,"OHKO",IF(I{row}>=0.5,"2HKO",IF(I{row}>=0.333,"3HKO",IF(I{row}>=0.25,"4HKO","5HKO+"))))')


def add_live_editor_sheets(
    workbook: Workbook,
    data: dict[str, Any],
    pokemon_by_id: dict[str, dict[str, Any]],
    ability_by_id: dict[str, dict[str, Any]],
    nature_by_name: dict[str, dict[str, Any]],
) -> None:
    """Add editable databases and a formula-driven damage calculator."""
    pokemon_rows = []
    for mon in sorted(data["pokemon"], key=lambda item: (item.get("number", 9999), item["id"])):
        types = list(mon.get("types", [])) + [""]
        abilities = list(mon.get("abilities", [])) + ["", ""]
        stats = mon["stats"]
        pokemon_rows.append([
            mon["id"], mon.get("rosterName") or mon["name"], types[0], types[1],
            stats["hp"], stats["attack"], stats["defense"], stats["spAttack"],
            stats["spDefense"], stats["speed"], abilities[0], abilities[1], abilities[2],
        ])
    append_sheet(workbook, "Pokemon Editor", [
        "Species ID", "Name", "Type 1", "Type 2", "Base HP", "Base Atk", "Base Def",
        "Base SpA", "Base SpD", "Base Spe", "Ability 1", "Ability 2", "Ability 3",
    ], pokemon_rows)

    evolution_rows = []
    for mon in sorted(data["pokemon"], key=lambda item: (item.get("number", 9999), item["id"])):
        seen_targets: set[str] = set()
        for evolution in mon.get("evolutions", []):
            method = str(evolution.get("method", ""))
            parameter = str(evolution.get("parameter", ""))
            target = evolution.get("target", "")
            digits = "".join(character for character in parameter if character.isdigit())
            if "LEVEL" in method and digits and int(digits) > 0:
                effective_level = int(digits)
            else:
                # Items, trades, friendship, move knowledge, party conditions,
                # and exported condition-driven Lv. 0 methods start at Lv. 40.
                effective_level = 40
            enabled = target not in seen_targets and not seen_targets
            evolution_rows.append([
                mon["id"], mon.get("rosterName") or mon["name"], method, target,
                effective_level, enabled, parameter,
            ])
            seen_targets.add(target)
    append_sheet(workbook, "Evolution Editor", [
        "Source Species ID", "Source", "Original method", "Target Species ID",
        "Effective level", "Enabled path?", "Original parameter",
    ], evolution_rows)

    learnset_rows = []
    move_name_by_id = {move["id"]: move["name"] for move in data["moves"]}
    for mon in sorted(data["pokemon"], key=lambda item: (item.get("number", 9999), item["id"])):
        for entry in mon.get("learnsets", {}).get("levelUp", []):
            move_id = entry["move"]
            learnset_rows.append([
                mon["id"], mon.get("rosterName") or mon["name"], int(entry["level"]),
                move_name_by_id.get(move_id, move_id), move_id,
            ])
    append_sheet(workbook, "Learnset Editor", [
        "Species ID", "Pokémon", "Level", "Move", "Move ID",
    ], learnset_rows)

    move_rows = [[
        move["id"], move["name"], move["type"], move["category"],
        move.get("power", 0), move.get("accuracy", 0), move.get("pp", 0),
    ] for move in sorted(data["moves"], key=lambda item: item["name"])]
    append_sheet(workbook, "Move Editor", [
        "Move ID", "Name", "Type", "Category", "Power", "Accuracy", "PP",
    ], move_rows)

    nature_rows = []
    neutral_base = {key: 100 for key in STAT_KEYS}
    for nature in sorted(data["natures"], key=lambda item: item["name"]):
        multipliers, notes = nature_stat_multipliers(
            nature["name"], neutral_base, nature_by_name,
            {"level": PLAYER_LEVEL, "full_hp": True, "met_level": PLAYER_LEVEL,
             "quirky_multiplier": 1.03},
        )
        nature_rows.append([
            nature["name"], multipliers["attack"], multipliers["defense"],
            multipliers["spAttack"], multipliers["spDefense"], multipliers["speed"],
            nature.get("description", ""), "; ".join(notes),
        ])
    nature_rows.insert(0, ["Neutral", 1, 1, 1, 1, 1, "No stat changes.", ""])
    append_sheet(workbook, "Nature Editor", [
        "Nature", "Atk x", "Def x", "SpA x", "SpD x", "Spe x", "Description", "Model notes",
    ], nature_rows)

    type_names = sorted(TYPE_CHART)
    chart_rows = []
    for attacking in type_names:
        chart_rows.append([attacking] + [TYPE_CHART.get(attacking, {}).get(defending, 1) for defending in type_names])
    append_sheet(workbook, "Type Chart", ["Attacking type"] + type_names, chart_rows)

    boss_rows = []
    for slot, config in enumerate(BOSS_TEAM, 1):
        species_id = norm_id(config["species"], "SPECIES_")
        mon = pokemon_by_id[species_id]
        moves = [norm_id(move, "MOVE_") for move in config.get("moves", [])]
        moves += [""] * (4 - len(moves))
        evs = complete_spread(config.get("evs"))
        boss_rows.append([
            slot, BOSS_NAME, species_id, int(config["level"]), config.get("nature", "Neutral"),
            config.get("ability") or (mon.get("abilities") or [""])[0],
            evs["hp"], evs["attack"], evs["defense"], evs["spAttack"],
            evs["spDefense"], evs["speed"], *moves[:4],
        ])
    append_sheet(workbook, "Boss Editor", [
        "Slot", "Boss", "Species ID", "Level", "Nature", "Ability",
        "HP EV", "Atk EV", "Def EV", "SpA EV", "SpD EV", "Spe EV",
        "Move 1", "Move 2", "Move 3", "Move 4",
    ], boss_rows)

    live = workbook.create_sheet("Live Calculator")
    live.append(["LIVE DAMAGE CALCULATOR", "Attacker", "", "Defender", ""])
    labels = [
        "Species ID", "Level", "Nature", "Ability", "Move ID",
        "HP EV", "Atk EV", "Def EV", "SpA EV", "SpD EV", "Spe EV",
    ]
    attacker_defaults = ["SPECIES_FLAAFFY", PLAYER_LEVEL, "Neutral", "", "MOVE_THUNDER_SHOCK", 0, 0, 0, 0, 0, 0]
    defender_default = norm_id(BOSS_TEAM[BOSS_TARGET_INDEX]["species"], "SPECIES_")
    defender_defaults = [defender_default, int(BOSS_TEAM[BOSS_TARGET_INDEX]["level"]),
                         BOSS_TEAM[BOSS_TARGET_INDEX].get("nature", "Neutral"),
                         BOSS_TEAM[BOSS_TARGET_INDEX].get("ability", ""), "", 0, 0, 0, 0, 0, 0]
    for label, attacker, defender in zip(labels, attacker_defaults, defender_defaults):
        live.append([label, attacker, "", label, defender])

    live.append([])
    live.append(["Calculated stat", "Attacker", "", "Calculated stat", "Defender"])
    stat_labels = ["HP", "Attack", "Defense", "Sp. Atk", "Sp. Def", "Speed"]
    base_columns = [5, 6, 7, 8, 9, 10]
    ev_rows = [7, 8, 9, 10, 11, 12]
    nature_columns = [None, 2, 3, 4, 5, 6]
    for index, (label, base_col, ev_row, nature_col) in enumerate(zip(stat_labels, base_columns, ev_rows, nature_columns), 15):
        live.cell(index, 1, label)
        live.cell(index, 4, label)
        for value_col, species_cell, level_cell, nature_cell, ev_cell in [
            (2, "$B$2", "$B$3", "$B$4", f"$B${ev_row}"),
            (5, "$E$2", "$E$3", "$E$4", f"$E${ev_row}"),
        ]:
            base = f'XLOOKUP({species_cell},\'Pokemon Editor\'!$A:$A,\'Pokemon Editor\'!${get_column_letter(base_col)}:${get_column_letter(base_col)},0)'
            if label == "HP":
                formula = f'=FLOOR(((2*{base}+31+FLOOR({ev_cell}/4))*{level_cell})/100)+{level_cell}+10'
            else:
                nature_mult = f'XLOOKUP({nature_cell},\'Nature Editor\'!$A:$A,\'Nature Editor\'!${get_column_letter(nature_col)}:${get_column_letter(nature_col)},1)'
                formula = f'=FLOOR((FLOOR(((2*{base}+31+FLOOR({ev_cell}/4))*{level_cell})/100)+5)*{nature_mult})'
            live.cell(index, value_col, formula)

    result_start = 22
    results = [
        ("Move name", '=XLOOKUP($B$6,\'Move Editor\'!$A:$A,\'Move Editor\'!$B:$B,"Unknown")'),
        ("Move type", '=XLOOKUP($B$6,\'Move Editor\'!$A:$A,\'Move Editor\'!$C:$C,"")'),
        ("Category", '=XLOOKUP($B$6,\'Move Editor\'!$A:$A,\'Move Editor\'!$D:$D,"")'),
        ("Power", '=XLOOKUP($B$6,\'Move Editor\'!$A:$A,\'Move Editor\'!$E:$E,0)'),
        ("Attack stat", '=IF(B24="Physical",B16,B18)'),
        ("Defense stat", '=IF(B24="Physical",E17,E19)'),
        ("STAB", '=IF(OR(B23=XLOOKUP($B$2,\'Pokemon Editor\'!$A:$A,\'Pokemon Editor\'!$C:$C,""),B23=XLOOKUP($B$2,\'Pokemon Editor\'!$A:$A,\'Pokemon Editor\'!$D:$D,"")),1.5,1)'),
        ("Type effectiveness", '=INDEX(\'Type Chart\'!$B:$S,MATCH(B23,\'Type Chart\'!$A:$A,0),MATCH(XLOOKUP($E$2,\'Pokemon Editor\'!$A:$A,\'Pokemon Editor\'!$C:$C,""),\'Type Chart\'!$B$1:$S$1,0))*IF(XLOOKUP($E$2,\'Pokemon Editor\'!$A:$A,\'Pokemon Editor\'!$D:$D,"")="",1,INDEX(\'Type Chart\'!$B:$S,MATCH(B23,\'Type Chart\'!$A:$A,0),MATCH(XLOOKUP($E$2,\'Pokemon Editor\'!$A:$A,\'Pokemon Editor\'!$D:$D,""),\'Type Chart\'!$B$1:$S$1,0)))'),
        ("Other multiplier", 1),
        ("Base damage", '=IF(OR(B24="Status",B25=0),0,FLOOR(FLOOR(FLOOR((2*$B$3/5+2)*B25*B26/B27)/50)+2))'),
        ("Minimum damage", '=FLOOR(B31*B28*B29*B30*0.85)'),
        ("Maximum damage", '=FLOOR(B31*B28*B29*B30)'),
        ("Min %", '=IF(E15=0,0,B32/E15)'),
        ("Max %", '=IF(E15=0,0,B33/E15)'),
    ]
    for offset, (label, value) in enumerate(results):
        live.cell(result_start + offset, 1, label)
        live.cell(result_start + offset, 2, value)
    live.cell(37, 1, "Notes")
    live.cell(37, 2, "Edit yellow input cells or any editor database. IVs are fixed at 31. Other multiplier is for contextual abilities, items, weather, terrain, or custom Nature effects not represented by raw stats.")
    for row in list(range(2, 13)):
        live.cell(row, 2).fill = PatternFill("solid", fgColor="FFF2CC")
        live.cell(row, 5).fill = PatternFill("solid", fgColor="FFF2CC")
    live.cell(30, 2).fill = PatternFill("solid", fgColor="FFF2CC")
    for row in (34, 35):
        live.cell(row, 2).number_format = "0.0%"
    live.column_dimensions["A"].width = 24
    live.column_dimensions["B"].width = 34
    live.column_dimensions["C"].width = 4
    live.column_dimensions["D"].width = 24
    live.column_dimensions["E"].width = 34


def upload_workbook_to_google_sheets(workbook: Workbook) -> str:
    """Replace this tool's tabs inside the configured native Google Sheet."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as exc:
        raise RuntimeError(
            "Google Sheets export needs: python3 -m pip install "
            "gspread google-auth"
        ) from exc

    credentials_path = Path(GOOGLE_SERVICE_ACCOUNT_FILE).expanduser()
    if not credentials_path.is_absolute():
        credentials_path = Path(__file__).resolve().parent / credentials_path
    if not credentials_path.exists():
        raise FileNotFoundError(
            f"Google credentials not found: {credentials_path}\n"
            "Download a service-account JSON key, save it with that name, and "
            "share the target Sheet with its client_email as an Editor."
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_file(
        str(credentials_path), scopes=scopes
    )
    client = gspread.authorize(credentials)

    def google_retry(operation, *args, **kwargs):
        """Retry transient Google API failures without hiding real errors."""
        last_error = None
        for attempt in range(5):
            try:
                return operation(*args, **kwargs)
            except gspread.exceptions.APIError as exc:
                last_error = exc
                status = getattr(getattr(exc, "response", None), "status_code", None)
                if status not in (429, 500, 502, 503, 504, None):
                    raise
                if attempt == 4:
                    raise
                time.sleep(2 ** attempt)
        raise last_error

    try:
        spreadsheet = google_retry(client.open_by_key, GOOGLE_SPREADSHEET_ID)
    except PermissionError as exc:
        raise PermissionError(
            "The Google service account cannot access the target spreadsheet. "
            "Open the Sheet's Share dialog and add this exact address as an "
            f"Editor: {credentials.service_account_email}"
        ) from exc

    generated_titles = [sheet.title for sheet in workbook.worksheets]
    staging_title = "_second_nature_refreshing"
    existing = {
        sheet.title: sheet for sheet in google_retry(spreadsheet.worksheets)
    }
    staging = existing.get(staging_title)
    if staging is None:
        staging = google_retry(
            spreadsheet.add_worksheet, staging_title, rows=2, cols=2
        )

    preserved_titles = (
        set() if RESET_GOOGLE_EDITOR_TABS
        else GOOGLE_EDITOR_TABS.intersection(existing)
    )
    # Schema migration: the evolution-aware Threshold Summary has an origin
    # species column. Replace the older snapshot once, then preserve it.
    if "Threshold Summary" in preserved_titles:
        old_threshold = existing["Threshold Summary"]
        if (old_threshold.col_count < 25
                or google_retry(old_threshold.acell, "Y1").value
                != "Origin Species ID"):
            preserved_titles.remove("Threshold Summary")
    # Calculated/report tabs are complete refreshes. Editor tabs are preserved
    # by default so hand-tuned stats and moves survive ordinary script reruns.
    for title in generated_titles:
        if title in preserved_titles:
            continue
        try:
            worksheet = google_retry(spreadsheet.worksheet, title)
            google_retry(spreadsheet.del_worksheet, worksheet)
        except gspread.WorksheetNotFound:
            pass

    requests: list[dict[str, Any]] = []
    preserved_threshold_source = None
    for source in workbook.worksheets:
        if source.title in preserved_titles:
            # Threshold is partly an editor (level, Nature, EVs, overrides) and
            # partly a calculated report. Refresh derived columns while leaving
            # move choice, level/Nature/ability/EVs, origin, and override intact.
            if source.title == "Threshold Summary":
                preserved_threshold_source = source
            continue
        values = [
            [google_cell_value(cell.value) for cell in row]
            for row in source.iter_rows(
                min_row=1, max_row=source.max_row,
                min_col=1, max_col=source.max_column,
            )
        ]
        row_count = max(2, len(values) + 5)
        column_count = max(1, max((len(row) for row in values), default=1))
        target = google_retry(
            spreadsheet.add_worksheet,
            source.title, rows=row_count, cols=column_count
        )
        google_retry(
            target.update,
            values=values,
            range_name="A1",
            value_input_option="USER_ENTERED",
        )
        sheet_id = target.id
        data_rows = max(1, len(values))

        requests.extend([
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": data_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": column_count,
                    },
                    "cell": {"userEnteredFormat": {
                        "verticalAlignment": "TOP",
                        "wrapStrategy": "WRAP",
                        "textFormat": {"fontFamily": "Arial", "fontSize": 10},
                    }},
                    "fields": "userEnteredFormat(verticalAlignment,wrapStrategy,textFormat)",
                }
            },
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": 1,
                        "startColumnIndex": 0,
                        "endColumnIndex": column_count,
                    },
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": .09, "green": .13, "blue": .20},
                        "textFormat": {
                            "bold": True,
                            "foregroundColor": {"red": 1, "green": 1, "blue": 1},
                        },
                        "horizontalAlignment": "LEFT",
                    }},
                    "fields": (
                        "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                    ),
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": sheet_id,
                        "gridProperties": {"frozenRowCount": 1},
                    },
                    "fields": "gridProperties.frozenRowCount",
                }
            },
            {
                "setBasicFilter": {"filter": {"range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": data_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": column_count,
                }}}
            },
            {
                "addBanding": {"bandedRange": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": data_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": column_count,
                    },
                    "rowProperties": {
                        "headerColor": {"red": .09, "green": .13, "blue": .20},
                        "firstBandColor": {"red": .91, "green": .95, "blue": .98},
                        "secondBandColor": {"red": 1, "green": 1, "blue": 1},
                    },
                }}
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id, "dimension": "ROWS",
                        "startIndex": 0, "endIndex": data_rows,
                    },
                    "properties": {"pixelSize": 24},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id, "dimension": "COLUMNS",
                        "startIndex": 0, "endIndex": column_count,
                    },
                    "properties": {"pixelSize": 125},
                    "fields": "pixelSize",
                }
            },
        ])

        headers = values[0] if values else []
        for column_index, header in enumerate(headers):
            header_text = str(header)
            if any(fragment in header_text for fragment in (
                "Notes", "notes", "Description", "description", "Available from",
                "Source", "Value",
            )):
                requests.append({"updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id, "dimension": "COLUMNS",
                        "startIndex": column_index, "endIndex": column_index + 1,
                    },
                    "properties": {"pixelSize": 420},
                    "fields": "pixelSize",
                }})
            if "%" in header_text and data_rows > 1:
                requests.extend([
                    {"repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1, "endRowIndex": data_rows,
                            "startColumnIndex": column_index,
                            "endColumnIndex": column_index + 1,
                        },
                        "cell": {"userEnteredFormat": {"numberFormat": {
                            "type": "PERCENT", "pattern": "0.0%",
                        }}},
                        "fields": "userEnteredFormat.numberFormat",
                    }},
                    {"addConditionalFormatRule": {
                        "index": 0,
                        "rule": {
                            "ranges": [{
                                "sheetId": sheet_id,
                                "startRowIndex": 1, "endRowIndex": data_rows,
                                "startColumnIndex": column_index,
                                "endColumnIndex": column_index + 1,
                            }],
                            "gradientRule": {
                                "minpoint": {
                                    "type": "MIN",
                                    "color": {"red": .39, "green": .75, "blue": .48},
                                },
                                "midpoint": {
                                    "type": "PERCENTILE", "value": "50",
                                    "color": {"red": 1, "green": .92, "blue": .52},
                                },
                                "maxpoint": {
                                    "type": "MAX",
                                    "color": {"red": .97, "green": .41, "blue": .42},
                                },
                            },
                        },
                    }},
                ])

    # Keep each API transaction comfortably below request-size limits.
    for start in range(0, len(requests), 80):
        google_retry(
            spreadsheet.batch_update,
            {"requests": requests[start:start + 80]},
        )

    # Refresh Threshold only after Move Comparison and the other dependencies
    # exist, so Sheets recalculates every best-move formula immediately.
    if preserved_threshold_source is not None:
        source = preserved_threshold_source
        target = existing[source.title]
        for start_col, end_col in ((1, 3), (5, 11), (21, 24)):
            formulas = [
                [google_cell_value(source.cell(row, column).value)
                 for column in range(start_col, end_col + 1)]
                for row in range(2, source.max_row + 1)
            ]
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            google_retry(
                target.update,
                values=formulas,
                range_name=f"{start_letter}2:{end_letter}{source.max_row}",
                value_input_option="USER_ENTERED",
            )

    google_retry(spreadsheet.del_worksheet, staging)
    try:
        default_sheet = spreadsheet.worksheet("Sheet1")
        if len(spreadsheet.worksheets()) > 1 and not default_sheet.get_all_values():
            google_retry(spreadsheet.del_worksheet, default_sheet)
    except gspread.WorksheetNotFound:
        pass
    return f"https://docs.google.com/spreadsheets/d/{GOOGLE_SPREADSHEET_ID}/edit"


def add_table(sheet, name: str) -> None:
    if sheet.max_row < 2 or sheet.max_column < 1:
        return
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)


def append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    return sheet


def pct(value: float, hp: int) -> float:
    return value / hp if hp else 0


def main() -> tuple[Path, str | None]:
    dex_path = find_dex_json()
    with dex_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)

    pokemon_by_id = {pokemon["id"]: pokemon for pokemon in data["pokemon"]}
    move_by_id = {move["id"]: move for move in data["moves"]}
    ability_by_id = {ability["id"]: ability for ability in data["abilities"]}
    nature_by_name = {nature["name"].lower(): nature for nature in data["natures"]}
    player_species, encounter_locations = collect_player_species(data, pokemon_by_id)

    bosses: list[tuple[Combatant, list[dict[str, Any]], list[str]]] = []
    for config in BOSS_TEAM:
        species_id = norm_id(config["species"], "SPECIES_")
        if species_id not in pokemon_by_id:
            raise KeyError(f"Boss species not found in dex: {species_id}")
        pokemon = pokemon_by_id[species_id]
        ability_id = norm_id(config.get("ability"), "ABILITY_") if config.get("ability") else (pokemon.get("abilities") or [None])[0]
        combatant, notes = make_combatant(
            pokemon, int(config["level"]), complete_spread(config.get("ivs"), 31),
            complete_spread(config.get("evs")), config.get("nature", "Neutral"),
            ability_id, config.get("item"), BOSS_STATUS, BOSS_AT_FULL_HP,
            nature_by_name, ability_by_id,
        )
        moves = legal_moves(pokemon, combatant.level, move_by_id, combatant.nature, config.get("moves"))
        bosses.append((combatant, moves, notes))

    if not (0 <= BOSS_TARGET_INDEX < len(bosses)) or not (0 <= BOSS_ATTACKER_INDEX < len(bosses)):
        raise IndexError("BOSS_TARGET_INDEX or BOSS_ATTACKER_INDEX is outside BOSS_TEAM")
    target, _, target_notes = bosses[BOSS_TARGET_INDEX]
    boss_attacker, boss_moves, attacker_notes = bosses[BOSS_ATTACKER_INDEX]

    offense_rows: list[list[Any]] = []
    defense_rows: list[list[Any]] = []
    threshold_rows: list[list[Any]] = []
    comparison_rows: list[list[Any]] = []
    all_matchup_rows: list[list[Any]] = []

    for species_id in player_species:
        pokemon = pokemon_by_id[species_id]
        candidate_records = []
        for nature_name in nature_choices(data["natures"]):
            moves = legal_moves(pokemon, PLAYER_LEVEL, move_by_id, nature_name)
            for ability_id in select_abilities(pokemon):
                player, stat_notes = make_combatant(
                    pokemon, PLAYER_LEVEL, PLAYER_IVS, PLAYER_EVS, nature_name,
                    ability_id, None, PLAYER_STATUS, PLAYER_AT_FULL_HP,
                    nature_by_name, ability_by_id, met_level=min(PLAYER_LEVEL, 5),
                )
                result = best_result(player, target, moves, PLAYER_FIRST_TURN)
                candidate_records.append((metric(result) if result else -1, player, moves, result, stat_notes))

        _, player, moves, result, stat_notes = max(candidate_records, key=lambda record: record[0])
        outgoing_pct = pct(result.average if result else 0, target.stats["hp"])
        offense_rows.append([
            player.name, ", ".join(sorted(encounter_locations[species_id])), player.level,
            player.nature, player.ability_name,
            result.move_name if result else "No damaging move",
            result.move_type if result else "", result.category if result else "",
            result.power if result else 0, result.accuracy if result else 0,
            result.minimum if result else 0, result.maximum if result else 0,
            round(result.average, 2) if result else 0,
            round(result.expected, 2) if result else 0,
            outgoing_pct, ko_label(result, target.stats["hp"]),
            "; ".join((*stat_notes, *(result.notes if result else ()))),
            player.species_id,
        ])

        incoming = best_result(boss_attacker, player, boss_moves, PLAYER_FIRST_TURN)
        incoming_pct = pct(incoming.average if incoming else 0, player.stats["hp"])
        defense_rows.append([
            player.name, player.nature, player.ability_name, player.stats["hp"],
            boss_attacker.name, incoming.move_name if incoming else "No damaging move",
            incoming.move_type if incoming else "", incoming.category if incoming else "",
            incoming.minimum if incoming else 0, incoming.maximum if incoming else 0,
            round(incoming.average, 2) if incoming else 0,
            round(incoming.expected, 2) if incoming else 0,
            incoming_pct, ko_label(incoming, player.stats["hp"]),
            "Yes" if player.stats["speed"] > boss_attacker.stats["speed"] else "No",
            "; ".join(incoming.notes if incoming else ()),
            player.species_id,
        ])

        score = outgoing_pct - incoming_pct
        rating = "Dominant" if outgoing_pct >= .50 and incoming_pct < .50 else (
            "Strong" if outgoing_pct >= .35 and incoming_pct < .75 else (
            "Viable" if outgoing_pct >= .25 and incoming_pct < 1 else (
            "Risky" if outgoing_pct >= .15 and incoming_pct < 1.25 else "Poor"
        )))
        threshold_rows.append([
            player.name, result.move_name if result else "None", outgoing_pct,
            incoming.move_name if incoming else "None", incoming_pct,
            player.stats["speed"], boss_attacker.stats["speed"],
            "Yes" if player.stats["speed"] > boss_attacker.stats["speed"] else "No",
            round(score, 4), rating, player.species_id, player.level, player.nature,
            player.ability_id or "", PLAYER_EVS["hp"], PLAYER_EVS["attack"],
            PLAYER_EVS["defense"], PLAYER_EVS["spAttack"], PLAYER_EVS["spDefense"],
            PLAYER_EVS["speed"], result.minimum if result else 0,
            result.maximum if result else 0, incoming.minimum if incoming else 0,
            incoming.maximum if incoming else 0, player.species_id, "",
        ])

        for move in moves:
            move_result = calculate_damage(player, target, move, PLAYER_FIRST_TURN)
            if not move_result:
                continue
            comparison_rows.append([
                player.name, player.nature, player.ability_name, move_result.move_name,
                move_result.move_type, move_result.category, move_result.power,
                move_result.accuracy, move_result.effectiveness,
                move_result.minimum, move_result.maximum, round(move_result.average, 2),
                round(move_result.expected, 2), pct(move_result.average, target.stats["hp"]),
                ko_label(move_result, target.stats["hp"]), "; ".join(move_result.notes),
                player.species_id,
            ])

        for boss, _, _ in bosses:
            matchup = best_result(player, boss, moves, PLAYER_FIRST_TURN)
            all_matchup_rows.append([
                player.name, boss.name, boss.level, matchup.move_name if matchup else "None",
                matchup.move_type if matchup else "", matchup.minimum if matchup else 0,
                matchup.maximum if matchup else 0, round(matchup.average, 2) if matchup else 0,
                pct(matchup.average if matchup else 0, boss.stats["hp"]),
                ko_label(matchup, boss.stats["hp"]),
                player.species_id, boss.species_id,
            ])

    # Reserve stable formula rows for each roster slot. Learnset Editor decides
    # which moves occupy them, so adding/removing/releveling moves is live.
    comparison_rows = []
    for species_id in player_species:
        pokemon = pokemon_by_id[species_id]
        for move_slot in range(1, 33):
            comparison_rows.append([
                pokemon.get("rosterName") or pokemon["name"], PLAYER_NATURE,
                (pokemon.get("abilities") or [""])[0], "", "", "", 0, 0,
                0, 0, 0, 0, 0, 0, "", "", species_id, move_slot,
            ])
    threshold_rows.sort(key=lambda row: row[8], reverse=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    scenario_rows = [
        ["Scenario", SCENARIO_NAME],
        ["Generated", datetime.now().isoformat(timespec="seconds")],
        ["Dex snapshot", str(dex_path)],
        ["Dex generatedAt", data.get("generatedAt", "Unknown")],
        ["Google Sheet", f"https://docs.google.com/spreadsheets/d/{GOOGLE_SPREADSHEET_ID}/edit"],
        ["Level cap / player level", PLAYER_LEVEL],
        ["Player Nature mode", f"{PLAYER_NATURE_MODE}: {PLAYER_NATURE}"],
        ["Player Ability mode", PLAYER_ABILITY_MODE],
        ["Move ranking metric", BEST_MOVE_METRIC],
        ["Weather", WEATHER or "None"],
        ["Terrain", TERRAIN or "None"],
        ["Included maps", ", ".join(ENCOUNTER_MAPS)],
        ["Included encounter methods", ", ".join(sorted(ALLOWED_ENCOUNTER_METHODS))],
        ["Excluded prerequisite moves", ", ".join(sorted(EXCLUDED_PREREQUISITE_MOVES)) or "None"],
        ["Player species/forms analyzed", len(player_species)],
        ["Boss", BOSS_NAME],
        ["Target", f"{target.name} Lv. {target.level}"],
        ["Boss attacker", f"{boss_attacker.name} Lv. {boss_attacker.level}"],
        ["Important", "Empty boss move lists use all legal level-up moves. Enter exact trainers.party moves before treating boss offense as final."],
    ]
    scenario = append_sheet(workbook, "Scenario", ["Setting", "Value"], scenario_rows)

    threshold = append_sheet(workbook, "Threshold Summary", [
        "Pokémon", "Best outgoing move", "Outgoing avg %", "Strongest incoming move",
        "Incoming avg %", "Player Spe", "Boss Spe", "Outspeeds?", "Net score", "Rating",
        "Species ID", "Level", "Nature", "Ability", "HP EV", "Atk EV", "Def EV",
        "SpA EV", "SpD EV", "Spe EV", "Outgoing min", "Outgoing max",
        "Incoming min", "Incoming max", "Origin Species ID", "Move override",
    ], threshold_rows)
    offense = append_sheet(workbook, "Player vs Boss", [
        "Pokémon", "Available from", "Level", "Nature", "Ability", "Best move", "Type",
        "Category", "Power", "Accuracy", "Min damage", "Max damage", "Average damage",
        "Expected damage", "Average % HP", "KO threshold", "Calculation notes", "Species ID",
    ], offense_rows)
    defense = append_sheet(workbook, "Boss vs Players", [
        "Player Pokémon", "Nature", "Ability", "Player HP", "Boss attacker", "Best boss move",
        "Type", "Category", "Min damage", "Max damage", "Average damage", "Expected damage",
        "Average % HP", "KO threshold", "Player outspeeds?", "Calculation notes", "Species ID",
    ], defense_rows)
    comparison = append_sheet(workbook, "Move Comparison", [
        "Pokémon", "Nature", "Ability", "Move", "Type", "Category", "Power", "Accuracy",
        "Effectiveness", "Min damage", "Max damage", "Average damage", "Expected damage",
        "Average % HP", "KO threshold", "Calculation notes", "Origin Species ID",
        "Move slot",
    ], comparison_rows)
    matchups = append_sheet(workbook, "All Boss Matchups", [
        "Player Pokémon", "Boss Pokémon", "Boss level", "Best move", "Type", "Min damage",
        "Max damage", "Average damage", "Average % HP", "KO threshold",
        "Player Species ID", "Boss Species ID",
    ], all_matchup_rows)

    pool_rows = [[pokemon_by_id[species_id]["name"], species_id,
                  ", ".join(sorted(encounter_locations[species_id]))]
                 for species_id in player_species]
    pool = append_sheet(workbook, "Encounter Pool", ["Pokémon", "Species ID", "Source"], pool_rows)

    nature_rows = [[nature["name"], nature.get("description", ""),
                    nature.get("statUp", ""), nature.get("statDown", ""),
                    "Secret" if nature.get("secret") else "Public"]
                   for nature in sorted(data["natures"], key=lambda item: item["name"])]
    nature_sheet = append_sheet(workbook, "Nature Reference", [
        "Nature", "Current description", "Stat up", "Stat down", "Availability",
    ], nature_rows)

    coverage_rows = [
        ["Core stats", "Supported", "Gen-style stat formula, IVs, EVs, level and 15% canonical Nature modifiers"],
        ["Type chart / STAB", "Supported", "Modern 18-type chart, dual typing, Adaptability"],
        ["Damage rolls", "Supported", "Sixteen 85–100 rolls with integer flooring"],
        ["Accuracy / expected damage", "Supported", "Expected damage includes accuracy; KO labels exclude miss chance"],
        ["Weather / terrain", "Partial", "Basic power modifiers only; grounded checks and move-specific exceptions are not modeled"],
        ["Abilities", "Partial", "Common offensive/defensive modifiers and immunities implemented; see source hooks"],
        ["Move flags", "Partial", "dex.json lacks contact/punch/bite/sound/slicing flags; curated lists and a contact heuristic are used"],
        ["Variable-power moves", "Not modeled", "Moves whose exported power is zero are excluded from best-move selection"],
        ["Fixed-damage moves", "Not modeled", "Seismic Toss, Dragon Rage, Endeavor and similar effects are excluded"],
        ["Multi-hit moves", "Not modeled", "Exported base power is treated as one hit; do not use those rows as final thresholds"],
        ["Critical hits", "Not modeled", "Standard damage rows represent ordinary non-critical hits"],
        ["Items", "Not modeled", "Held-item damage modifiers and healing are reserved for the next pass"],
        ["Custom Natures", "Partial", "Direct stat/damage Natures are modeled where context is available; utility/healing effects remain descriptive"],
        ["AI and switching", "Not modeled", "Workbook compares damage thresholds, not optimal battle lines"],
        ["Boss trainer source", "Manual draft", "Current source snapshot still has vanilla Roxanne; edit BOSS_TEAM at top of script"],
    ]
    coverage = append_sheet(workbook, "Coverage", ["Mechanic", "Status", "Notes"], coverage_rows)

    add_live_editor_sheets(
        workbook, data, pokemon_by_id, ability_by_id, nature_by_name
    )
    make_analysis_tabs_live(threshold, offense, defense, comparison, matchups)

    for sheet, name in [
        (threshold, "ThresholdTable"), (offense, "PlayerOffenseTable"),
        (defense, "BossOffenseTable"), (comparison, "MoveComparisonTable"),
        (matchups, "AllMatchupsTable"), (pool, "EncounterPoolTable"),
        (nature_sheet, "NatureReferenceTable"), (coverage, "CoverageTable"),
    ]:
        add_table(sheet, name)

    # Percentage number formats and scan-friendly conditional formatting.
    for sheet, columns in [
        (threshold, [3, 5]), (offense, [15]), (defense, [13]),
        (comparison, [14]), (matchups, [9]),
    ]:
        for column in columns:
            for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=sheet.max_row):
                for entry in cell:
                    entry.number_format = "0.0%"
            rng = f"{get_column_letter(column)}2:{get_column_letter(column)}{sheet.max_row}"
            sheet.conditional_formatting.add(
                rng, ColorScaleRule(start_type="min", start_color="63BE7B",
                                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                    end_type="max", end_color="F8696B")
            )

    # Reverse incoming-damage coloring so low incoming damage is green.
    threshold.conditional_formatting.add(
        f"E2:E{threshold.max_row}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B"),
    )

    style_workbook(workbook)
    scenario.column_dimensions["B"].width = 100
    coverage.column_dimensions["C"].width = 100
    nature_sheet.column_dimensions["B"].width = 100
    offense.column_dimensions["Q"].width = 55
    defense.column_dimensions["P"].width = 55
    comparison.column_dimensions["P"].width = 55

    # Compact chart: strongest average outgoing percentages at a glance.
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = f"Top outgoing damage into {target.name}"
    chart.y_axis.title = "Pokémon"
    chart.x_axis.title = "Average target HP"
    top_n = min(16, threshold.max_row)
    data_ref = Reference(threshold, min_col=3, min_row=1, max_row=top_n)
    cats_ref = Reference(threshold, min_col=1, min_row=2, max_row=top_n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 14
    threshold.add_chart(chart, "L2")

    output_path = Path(OUTPUT_FILE).expanduser()
    if not output_path.is_absolute():
        output_path = Path(__file__).resolve().parent / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    google_url = None
    if GOOGLE_SHEETS_ENABLED:
        google_url = upload_workbook_to_google_sheets(workbook)
    return output_path, google_url


if __name__ == "__main__":
    try:
        generated, google_sheet = main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
    print(f"Generated {generated}")
    if google_sheet:
        print(f"Updated {google_sheet}")
